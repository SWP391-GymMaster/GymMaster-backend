"""Dien Template0_AI Usage Report.

Cot 'Student's Validation/Modification' va 'Risks/Limitations Observed' la cho
THAY CHAM — thay muon thay sinh vien KIEM CHUNG lai AI, khong chep mu. Nen moi
dong duoi day deu la viec CO THAT da xay ra, kem bang chung kiem chung duoc.

Dung: python fill_ai_usage.py <template.xlsx> <out.xlsx>
"""
import argparse
import shutil
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thieu openpyxl. Chay: uv run --with openpyxl python ' + __file__)

# No | SDLC Phase | Task | AI Tool | AI Output | Validation/Modification |
# Evidence | Quantitative Measure | Value Added | Risks/Limitations
ROWS = [
    ('Requirement', 'Dong bo spec kit theo code that', 'Claude Code (Opus 4.8)',
     'Doc ~12k dong code backend roi viet lai 10 feature spec + 16 doc danh so',
     'Doi chieu tung spec voi code that: sua PACKAGE_PT_REQUIRED (409) thay vi '
     'NO_PT_PACKAGE (422), MaxPerDay=2, bo phan photo/Azure Blob chua implement',
     'commit 38fd876 (27 file, 664+/662-)', '26 file spec dong bo', 5,
     'AI de tin spec cu la dung. Phai bat doc code truoc, khong doc spec truoc.'),

    ('Design', 'Phat hien god node de gom feature', 'graphify (AST tree-sitter)',
     'Do thi 1229 node / 3399 edge; chi AuthServiceResult la god node degree 185',
     'GREP LAI 38 file de xac minh truoc khi tin -> dung: kieu nay dung o MOI '
     'feature, khong rieng auth -> doi ten thanh ServiceResult',
     'commit 9fa30f9', '24 -> 49 quan he ERD sau khi sua FK alias', 5,
     'AI chi ra dung nhung neu khong grep xac minh thi da gom AuthServiceResult '
     'vao Features/Auth va lam hong kien truc.'),

    ('Coding', 'Refactor 91 file sang feature-based', 'Claude Code (Opus 4.8)',
     'Di chuyen 91 file, doi namespace, sua ~190 dong using',
     'Chay build + 71 test TRUOC va SAU: deu 0 loi, 71/71 pass. Chay app that: '
     'login 200 + JWT, 5 endpoint deu 200. CS0118 (namespace CheckIn dung ten '
     'entity) chi lo ra luc build -> doi thanh CheckIns',
     'commit 9fa30f9 (106 file)', 'build 0 loi; test 71/71; 5/5 endpoint 200', 5,
     'Loi namespace dung ten entity KHONG the phat hien bang doc code, phai build.'),

    ('Testing', 'Them CI GitHub Actions', 'Claude Code (Opus 4.8)',
     'Workflow build + test + quet lo hong cho ca 2 repo',
     'Phat hien `dotnet list --vulnerable` LUON tra exit 0 ke ca khi co lo hong '
     '-> tin exit code la CI bao xanh GIA. Doi sang grep output. TEST 2 CHIEU '
     'bang cach ha Microsoft.OpenApi ve 2.0.0: co lo hong -> chan dung; da va -> cho qua',
     'commit 1767e70; CI run 29517691321 (success)', 'CI xanh ca 2 repo', 5,
     'AI viet CI trong dung nhung im lang sai. Khong test 2 chieu thi cong bao '
     'mat vo dung mai mai.'),

    ('Testing', 'Chay full test toan du an', 'Claude Code (Opus 4.8)',
     'Ket luan ban dau: "UI khong doi, anh chup van dung"',
     'SAI. Kiem tra lai: git ls-files = 0 anh duoc track. Git bao "0 thay doi" '
     'khong phai vi anh giong nhau ma vi .gitignore co luat /docs/ chan sach. '
     'Da chup lai 43 anh that',
     'commit b9fc6dd', '27 -> 43 anh; 43/43 test pass', 3,
     'AI doc nham tin hieu git va suyt ket luan sai. Phai hoi "vi sao 0 thay doi" '
     'chu khong nhan ket qua.'),

    ('Documentation', 'Sinh 30 bang use case cho RDS', 'Claude Code (Opus 4.8)',
     '30 bang 15 dong tu 03_SRS_USE_CASES.md + specs/00X/spec.md',
     'DOC NGUOC file .docx ra kiem tra -> gan nhu MOI O DEU RONG. Nguyen nhan: '
     'o gop trong Word dung chung mot <w:tc>, ghi cells[1] roi xoa cells[2:] la '
     'tu xoa mat chu vua ghi. Them uniq_cells() loc theo id(_tc)',
     'commit a9e4e63', '30 bang; 0/32 -> 31/32 co du lieu', 4,
     'Neu khong doc nguoc file ra kiem tra thi da nop mot RDS trang tron 30 bang.'),

    ('Documentation', 'Loc Business Rules cho tung use case', 'Claude Code (Opus 4.8)',
     'Loc FR theo tu khoa trong ten UC',
     'HONG: "Login" chi khop FR-RBAC-03 (chang lien quan), con FR-AUTH-02 (dung '
     'nghia login) bi loai vi noi dung viet "gui email + mat khau dung", khong co '
     'chu "login". Tu thua 16 FR thanh SAI 1 FR -> BO cach loc, quay lai liet ke '
     'du + danh dau [CAN CAT BOT] de nguoi doc tu cat',
     'commit a9e4e63', '16 FR -> 1 FR sai -> quay lai 16 FR', 2,
     'Viec can PHAN DOAN (FR nao ap dung cho UC nao) thi AI khong lam duoc. '
     'Liet ke thua thi nhin thay ma cat; loc sai thi khong ai biet.'),

    ('Documentation', 'Kiem tra template truoc khi nop', 'Claude Code (Opus 4.8)',
     'Dien RDS/SDS theo template cua thay',
     'Quet toan file phat hien template CHUA SAN du an mau cua NGUOI KHAC: '
     'Patron 114 lan, Payroll 23, Cafeteria 12, GAMS o chinh tieu de, anh co chu '
     '"Teacher". Da xoa sach 11/11 tu khoa',
     'commit 270b069', '11 tu khoa mau -> 0', 4,
     'AI dien template nhung khong tu xoa mau. Nop nguyen la nop nham du an nha an.'),

    ('Deployment', 'Va lo hong bao mat', 'Claude Code (Opus 4.8) + NuGet audit',
     'NU1903: Microsoft.OpenApi 2.0.0 dinh CVE-2026-49451 (CVSS 7.5 High)',
     'Kiem tra: nang Microsoft.AspNetCore.OpenApi len 10.0.10 KHONG sua duoc — ban '
     'moi nhat van khai bao phu thuoc 2.0.0. Phai ghim truc tiep 2.7.5. Xac minh: '
     'build 0 warning, test 71/71, /openapi/v1.json van sinh 90KB/69 endpoint',
     'commit 5c610c2', 'build 2 -> 0 warning; 71/71 test', 5,
     'Rui ro that gan 0 (app tu SINH OpenAPI, khong PARSE doc la) nhung van va de '
     'qua security scan. AI de de xuat "nang package" ma khong kiem tra co sua duoc khong.'),
]

HDR_ROW = 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('template')
    ap.add_argument('out')
    a = ap.parse_args()

    shutil.copy(a.template, a.out)
    wb = load_workbook(a.out)

    # --- Overview
    ws = wb['0.Overview']
    known = {
        'Subject Code': 'SWP391',
        'Subject Name': 'Application Development Project',
        'Project Title': 'GymMaster — Gym Management Web System',
        'Class Code': '[CAN BO SUNG]',
        'Semester': '[CAN BO SUNG]',
        'Lecturer Name': '[CAN BO SUNG]',
        'Group Code': '[CAN BO SUNG]',
    }
    for row in ws.iter_rows(max_row=12):
        for c in row:
            v = str(c.value).strip() if c.value else ''
            if v in known:
                ws.cell(row=c.row, column=c.column + 1, value=known[v])

    # danh sach sinh vien — chi biet nick git, khong biet MSSV
    members = [
        ('anhdaijka', 'Frontend (41/47 man)'),
        ('BanhMiChao', 'Backend Auth + FE Billing'),
        ('Loc-LX', 'Backend Billing/Nutrition/Members/Trainers/Users/Account'),
        ('vandam2005', 'Backend Training/Dashboard'),
        ('Minhdicodedao', 'Backend CheckIns'),
    ]
    hdr_r = None
    for row in ws.iter_rows(max_row=16):
        for c in row:
            if str(c.value).strip() == 'No':
                hdr_r = c.row
                break
        if hdr_r:
            break
    if hdr_r:
        for i, (nick, role) in enumerate(members, 1):
            r = hdr_r + i
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value='[CAN BO SUNG]')      # StudentCode
            ws.cell(row=r, column=3, value=nick)                  # StudentName
            ws.cell(row=r, column=4, value=role)                  # Role In Group
            ws.cell(row=r, column=5, value='Claude Code, GitHub Copilot')

    # --- Week sheets
    for name in wb.sheetnames:
        if not name[0].isdigit() or 'Overview' in name:
            continue
        w = wb[name]
        w.delete_rows(HDR_ROW + 1, max(w.max_row - HDR_ROW, 1))
        for i, r in enumerate(ROWS, 1):
            rr = HDR_ROW + i
            w.cell(row=rr, column=1, value=i)
            for j, v in enumerate(r, 2):
                w.cell(row=rr, column=j, value=v)
        break        # chi dien sheet tuan dau lam mau

    wb.save(a.out)
    print('Da dien AI Usage Report -> {}'.format(a.out))
    print('  {} dong log, tat ca deu la viec CO THAT kem link commit'.format(len(ROWS)))
    print('  CAN BO SUNG: Class Code, Semester, Lecturer Name, Group Code, MSSV tung nguoi')


if __name__ == '__main__':
    main()
