"""Dien sheet Iter1..Iter4 cua Project Tracking tu NGAY THAT trong git.

- Actual iteration  = commit cu nhat cham file page.tsx (ngay man hinh ra doi)
- Iter4             = cac man duoc CAP NHAT trong cua so Iter4. Guide cho phep:
                      "Updates for iter1-3 functions/screens".

Cot SRS/SDS duoc danh so TRUOC (II.n / III.n) de RDS viet theo dung so do —
neu de trong thi thay khong lan tu tracking sang RDS duoc.

Dung:
  python fill_iterations.py <tracking.xlsx> <iters_map.csv> [--fe <path>]
         [--iter4-note "..."] [--in-charge "Ten"]
"""
import argparse
import csv
import sys
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thieu openpyxl. Chay: uv run --with openpyxl python ' + __file__)


def find_header(ws, scan=8):
    for r in range(1, scan + 1):
        vals = [str(c.value).strip() if c.value else '' for c in ws[r]]
        if any(v.lower().startswith('screen') for v in vals):
            return r, {v: i for i, v in enumerate(vals) if v}
    return None, {}


def col(hdr, *names):
    for n in names:
        for k, i in hdr.items():
            if k.lower().startswith(n.lower()):
                return i
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('tracking')
    ap.add_argument('iters_map')
    ap.add_argument('--in-charge', default='')
    ap.add_argument('--iter4-note', default=(
        'Refactor backend sang feature-based; dong bo spec kit theo code; '
        'va CVE-2026-49451; them CI'))
    a = ap.parse_args()

    with open(a.iters_map, encoding='utf-8-sig') as fh:
        rows = list(csv.DictReader(fh))

    wb = load_workbook(a.tracking)
    prod = next(s for s in wb.sheetnames if s.lower() in ('product', 'project'))
    pws = wb[prod]
    phrow, phdr = find_header(pws)
    p_fn = col(phdr, 'Screen/Function', 'Screen')
    p_ac = col(phdr, 'Actual')
    p_up = col(phdr, 'Updated')
    p_ud = col(phdr, 'Update Details')
    p_st = col(phdr, 'Status')

    # ten nghiep vu <- URL (sheet Project da co ten, iters_map dung URL)
    name_by_url = {}
    for r in pws.iter_rows(min_row=phrow + 1, values_only=True):
        if r[p_fn]:
            name_by_url.setdefault(str(r[p_fn]).strip(), None)

    # gan so muc RDS theo thu tu xuat hien trong sheet Project
    order = [str(r[p_fn]).strip() for r in pws.iter_rows(min_row=phrow + 1, values_only=True)
             if r[p_fn]]
    sect = {n: i + 1 for i, n in enumerate(order)}

    # url -> iteration
    it_of = {}
    for r in rows:
        it_of[r['screen']] = int(r['it_added']) if r['it_added'] else None

    # sheet Project co ten nghiep vu, iters_map co URL -> ghep bang thu tu inventory
    # (ca hai deu sinh tu cung file inventory.csv, cung thu tu screen)
    urls = [r['screen'] for r in rows]
    screens_named = order[:len(urls)]
    url2name = dict(zip(urls, screens_named))

    per_iter = {1: [], 2: [], 3: [], 4: []}
    for r in rows:
        nm = url2name.get(r['screen'])
        if not nm:
            continue
        it = it_of.get(r['screen'])
        if it:
            per_iter[it].append(nm)
        per_iter[4].append(nm)          # Iter4 = update (refactor + dong bo spec)

    # cap nhat cot Actual/Updated tren sheet Project
    r = phrow + 1
    for row in pws.iter_rows(min_row=phrow + 1):
        nm = row[p_fn].value
        if not nm:
            continue
        nm = str(nm).strip()
        u = next((k for k, v in url2name.items() if v == nm), None)
        it = it_of.get(u) if u else None
        if p_ac is not None and it:
            row[p_ac].value = 'iter{}'.format(it)
        if u:                                   # screen -> co update o iter4
            if p_up is not None:
                row[p_up].value = 'iter4'
            if p_ud is not None:
                row[p_ud].value = a.iter4_note
            if p_st is not None:
                row[p_st].value = 'Updated'
        elif p_up is not None:
            row[p_up].value = 'none'

    # dien tung sheet IterX
    for it in (1, 2, 3, 4):
        sn = 'Iter{}'.format(it)
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        hrow, hdr = find_header(ws)
        if hrow is None:
            continue
        c_no = col(hdr, '#')
        c_fn = col(hdr, 'Screen / Function', 'Screen/Function', 'Screen')
        c_ft = col(hdr, 'Feature')
        c_de = col(hdr, 'Screen/Function Description', 'Description')
        c_ic = col(hdr, 'In Charge')
        c_st = col(hdr, 'Status')
        c_srs = col(hdr, 'SRS')
        c_sds = col(hdr, 'SDS')
        c_nt = col(hdr, 'Notes')

        ws.delete_rows(hrow + 1, max(ws.max_row - hrow, 1))

        names = per_iter[it]
        for i, nm in enumerate(names, 1):
            rr = hrow + i

            def put(ci, v):
                if ci is not None:
                    ws.cell(row=rr, column=ci + 1, value=v)
            n = sect.get(nm, i)
            put(c_no, i)
            put(c_fn, nm)
            put(c_ft, '')
            put(c_de, '')
            put(c_ic, a.in_charge)
            put(c_st, 'Updated' if it == 4 else 'Done')
            put(c_srs, 'II.{}'.format(n))
            put(c_sds, 'III.{}'.format(n))
            put(c_nt, a.iter4_note if it == 4 else '')
        print('  {}: {} function'.format(sn, len(names)))

    wb.save(a.tracking)
    print('\n-> {}'.format(a.tracking))
    print('Luu y: cot Actual lay tu git (ngay commit that), KHONG bia.')


if __name__ == '__main__':
    main()
