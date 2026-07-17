"""Dien sheet Project cua Template1_Project Tracking.xlsx tu inventory.csv.

Don vi dong = SCREEN (man hinh) + NON-UI FUNCTION, khong phai tung REST endpoint.
Thay cham "LOC x Quality" theo function nghiep vu (60/120/240, can >=720 ca du an),
nen 89 endpoint tho khong phai la 89 dong — chung nam trong RDS muc III cua tung man.

Dung: python fill_tracking.py <inventory.csv> <template.xlsx> <out.xlsx> [--in-charge "Ten"]
"""
import argparse
import csv
import re
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thieu openpyxl. Chay: uv run --with openpyxl python ' + __file__)

# Ten nghiep vu cho tung route — lay tu docs/design/08_ROUTE_MAP_NAVIGATION.md
# cua frontend (muc "Admin/Staff/PT/Member navigation").
NAMES = {
    '/': 'Landing Page', '/ (landing)': 'Landing Page',
    '/login': 'User Login', '/signup': 'User Register',
    '/forgot-password': 'Forgot Password', '/reset-password': 'Reset Password (OTP)',
    '/change-password': 'Change Password', '/welcome': 'Welcome Page', '/about': 'About Page',
    '/admin/dashboard': 'Admin Dashboard', '/admin/users': 'Manage User Accounts',
    '/admin/staff': 'Manage Staff Accounts', '/admin/members': 'Manage Members',
    '/admin/members/[id]': 'Member 360 Profile (Admin)', '/admin/trainers': 'Manage PTs',
    '/admin/packages': 'Manage Packages', '/admin/memberships': 'Manage Memberships',
    '/admin/payments': 'Payments (Admin)', '/admin/assignments': 'PT Assignment',
    '/admin/audit-logs': 'Audit Logs',
    '/admin/profile': 'Admin Profile',
    '/staff/dashboard': 'Staff Dashboard', '/staff/members': 'Member Search',
    '/staff/members/[id]': 'Member 360 Profile (Staff)', '/staff/sell-package': 'Sell Package',
    '/staff/renew-package': 'Renew Package', '/staff/check-in': 'Front-desk Check-in',
    '/staff/payments': 'Payments (Staff)', '/staff/profile': 'Staff Profile',
    '/pt/dashboard': 'PT Dashboard', '/pt/members': 'Assigned Members',
    '/pt/members/[id]': 'Member 360 Profile (PT)', '/pt/members/[id]/workout': 'Workout Plan (PT)',
    '/pt/members/[id]/notes': 'Trainer Notes', '/pt/members/[id]/progress': 'Member Progress (PT)',
    '/pt/check-in': 'PT Check-in', '/pt/profile': 'PT Profile',
    '/member/dashboard': 'Member Dashboard', '/member/membership': 'Buy / Renew Membership',
    '/member/membership/vnpay-return': 'VNPay Return', '/member/workout': 'My Workout Plan',
    '/member/notes': 'My Trainer Notes', '/member/progress': 'My Progress',
    '/member/nutrition/meal-journal': 'Meal Journal (+ AI Scan)',
    '/member/nutrition/summary': 'Calorie Summary',
    '/member/profile': 'My Profile', '/member/profile/edit': 'Edit My Profile',
}

# Function khong co giao dien — RDS muc I.2.4 hoi dung nhung cai nay.
NON_UI = [
    ('VNPay IPN Handler', 'Billing', 'System',
     'VNPay goi server-to-server bao ket qua thanh toan; xac thuc chu ky HMAC-SHA512 '
     'roi cap nhat Payment/Membership. Khong co UI.'),
    ('Membership Auto-Cancel', 'Billing', 'System',
     'MembershipLifecycle tu huy membership o trang thai PendingPayment qua 30 phut '
     '(PendingPaymentTtl). Chay khi co request cham vao membership.'),
    ('Membership Lazy Expire', 'Billing', 'System',
     'Membership Active qua EndDate duoc chuyen sang Expired khi doc (lazy), '
     'moc thoi gian theo AppClock GMT+7.'),
]


def find_header(ws, scan=8):
    for r in range(1, scan + 1):
        vals = [str(c.value).strip() if c.value else '' for c in ws[r]]
        if any(v.lower().startswith('screen') for v in vals):
            return r, {v: i for i, v in enumerate(vals) if v}
    sys.exit('Khong tim thay dong tieu de trong sheet')


def col(hdr, *names):
    for n in names:
        for k, i in hdr.items():
            if k.lower().startswith(n.lower()):
                return i
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('inventory')
    ap.add_argument('template')
    ap.add_argument('out')
    ap.add_argument('--in-charge', default='')
    a = ap.parse_args()

    with open(a.inventory, encoding='utf-8-sig') as fh:
        inv = list(csv.DictReader(fh))
    screens = [r for r in inv if r['kind'] == 'Screen']

    shutil.copy(a.template, a.out)           # luon dien vao BAN SAO
    wb = load_workbook(a.out)
    sheet = next((s for s in wb.sheetnames if s.lower() in ('product', 'project')), None)
    ws = wb[sheet]
    hrow, hdr = find_header(ws)

    c_no = col(hdr, '#')
    c_fn = col(hdr, 'Screen/Function', 'Screen')
    c_ft = col(hdr, 'Feature')
    c_ac = col(hdr, 'Actor')
    c_de = col(hdr, 'Screen/Function Description', 'Description')
    c_ic = col(hdr, 'In Charge')
    c_st = col(hdr, 'Status')

    # xoa du lieu mau
    ws.delete_rows(hrow + 1, ws.max_row - hrow)

    rows = []
    unknown = []
    for s in screens:
        name = NAMES.get(s['name'])
        if not name:
            unknown.append(s['name'])
            name = s['name']
        rows.append((name, s['feature'].capitalize(), s['actor'],
                     'Man hinh {} ({})'.format(s['name'], s['desc'])))
    for n, f, ac, d in NON_UI:
        rows.append((n, f, ac, d))

    r = hrow + 1
    for i, (name, feat, actor, desc) in enumerate(rows, 1):
        def put(ci, v):
            if ci is not None:
                ws.cell(row=r, column=ci + 1, value=v)
        put(c_no, i)
        put(c_fn, name)
        put(c_ft, feat)
        put(c_ac, actor)
        put(c_de, desc)
        put(c_ic, a.in_charge)
        put(c_st, 'Done')
        r += 1

    wb.save(a.out)
    print('sheet {}: da ghi {} dong ({} screen + {} non-UI)'.format(
        sheet, len(rows), len(screens), len(NON_UI)))
    print('-> {}'.format(a.out))
    if unknown:
        print('\nCHUA CO TEN NGHIEP VU (dang de nguyen URL, can dat ten tay):')
        for u in unknown:
            print('  ' + u)


if __name__ == '__main__':
    main()
