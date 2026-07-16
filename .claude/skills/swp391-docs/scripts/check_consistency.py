"""Kiem tra tinh nhat quan giua cac file doc SWP391 truoc khi nop.

Bat 4 loi hay gap nhat — nhung loi lam ca goi tai lieu roi rac:
  1. Ten function trong Issues Report khong co trong Project Tracking
  2. Sheet IterX co function khong co trong sheet Product
  3. Cot SRS/SDS bo trong (thay khong lan duoc tu tracking sang RDS)
  4. Endpoint co that trong code nhung thieu trong sheet Product

Dung: python check_consistency.py <tracking.xlsx> [issues.xlsx] [--inventory inv.csv]
"""
import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thieu openpyxl. Chay: uv run --with openpyxl python ' + __file__)


def header_map(ws, row=1, scan=6):
    """Tim dong tieu de (template co vai dong ghi chu o tren)."""
    for r in range(1, scan + 1):
        vals = [(c.value or '') for c in ws[r]]
        joined = ' '.join(str(v) for v in vals).lower()
        if 'screen' in joined or 'title' in joined:
            return r, {str(v).strip(): i for i, v in enumerate(vals) if v}
    return row, {}


def col(hdr, *names):
    for n in names:
        for k, i in hdr.items():
            if k.lower().startswith(n.lower()):
                return i
    return None


def read_rows(ws, hrow):
    for r in ws.iter_rows(min_row=hrow + 1):
        vals = [(c.value if c.value is not None else '') for c in r]
        if any(str(v).strip() for v in vals):
            yield vals


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('tracking')
    ap.add_argument('issues', nargs='?')
    ap.add_argument('--inventory')
    a = ap.parse_args()

    problems = []

    wb = load_workbook(a.tracking, data_only=True)
    # Student Guides goi la sheet "Product", nhung file .xlsx that dat ten
    # la "Project". Chap nhan ca hai.
    sheet = next((s for s in wb.sheetnames if s.lower() in ('product', 'project')), None)
    if not sheet:
        sys.exit('Khong thay sheet Product/Project trong {} (co: {})'
                 .format(a.tracking, wb.sheetnames))

    ws = wb[sheet]
    hrow, hdr = header_map(ws)
    ci = col(hdr, 'Screen/Function', 'Screen')
    if ci is None:
        sys.exit('Khong tim thay cot Screen/Function trong sheet Product')
    product = set()
    for vals in read_rows(ws, hrow):
        v = str(vals[ci]).strip()
        if v:
            product.add(v)
    print('sheet {}: {} function'.format(sheet, len(product)))

    # 2 + 3. sheet IterX
    for name in wb.sheetnames:
        if not name.lower().startswith('iter'):
            continue
        w = wb[name]
        hr, h = header_map(w)
        c_fn = col(h, 'Screen / Function', 'Screen/Function', 'Screen')
        c_srs, c_sds = col(h, 'SRS'), col(h, 'SDS')
        if c_fn is None:
            continue
        for vals in read_rows(w, hr):
            fn = str(vals[c_fn]).strip()
            if not fn:
                continue
            if fn not in product:
                problems.append('[{}] "{}" khong co trong sheet {}'.format(name, fn, sheet))
            for lbl, ci2 in (('SRS', c_srs), ('SDS', c_sds)):
                if ci2 is not None and not str(vals[ci2] if ci2 < len(vals) else '').strip():
                    problems.append('[{}] "{}" bo trong cot {} (thay khong lan duoc sang RDS)'
                                    .format(name, fn, lbl))

    # 1. Issues Report
    if a.issues and Path(a.issues).exists():
        iw = load_workbook(a.issues, data_only=True)
        s = iw[iw.sheetnames[0]]
        hr, h = header_map(s)
        c_fn = col(h, 'Functions/Screens', 'Function')
        if c_fn is not None:
            for vals in read_rows(s, hr):
                fn = str(vals[c_fn] if c_fn < len(vals) else '').strip()
                if fn and fn not in product:
                    problems.append('[Issues] "{}" khong khop ten nao trong sheet {}'
                                    .format(fn, sheet))

    # 4. code vs tracking
    if a.inventory and Path(a.inventory).exists():
        with open(a.inventory, encoding='utf-8-sig') as fh:
            inv = [r for r in csv.DictReader(fh)]
        missing = [r['name'] for r in inv if r['name'] not in product]
        if missing:
            problems.append('[Code] {} endpoint/screen co that nhung THIEU trong Product, vd: {}'
                            .format(len(missing), ', '.join(missing[:5])))

    print()
    if problems:
        print('CO {} VAN DE:'.format(len(problems)))
        for p in problems:
            print('  - ' + p)
        sys.exit(1)
    print('OK - cac file khop nhau.')


if __name__ == '__main__':
    main()
