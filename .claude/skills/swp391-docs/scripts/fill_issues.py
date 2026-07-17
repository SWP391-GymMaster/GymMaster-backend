"""Dien Template4_Issues Report tu GIT LOG THAT cua ca 2 repo.

Khong bia issue: moi dong = mot commit co that, co ngay + tac gia + link GitHub
kiem chung duoc. Conventional commit prefix -> Labels:
  feat -> Task/WP · fix -> Defect · docs/chore/test/refactor -> Task

Cot Functions/Screens PHAI khop TUNG KY TU voi sheet Project cua Project
Tracking, neu khong check_consistency.py se bao lech.

Dung: python fill_issues.py <template.xlsx> <out.xlsx> <tracking.xlsx>
"""
import argparse
import re
import shutil
import subprocess
import sys
from datetime import date, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thieu openpyxl. Chay: uv run --with openpyxl python ' + __file__)

REPOS = [
    ('backend', r'D:\GymMaster\GymMaster-backend',
     'https://github.com/SWP391-GymMaster/GymMaster-backend'),
    ('frontend', r'D:\GymMaster\GymMaster-frontend',
     'https://github.com/SWP391-GymMaster/GymMaster-frontend'),
]
START = date(2026, 5, 31)

# tu khoa trong commit -> ten function trong sheet Project
KEY2FN = [
    ('vnpay', 'Buy / Renew Membership'), ('payment', 'Payments (Admin)'),
    ('billing', 'Buy / Renew Membership'), ('membership', 'Manage Memberships'),
    ('package', 'Manage Packages'), ('sell', 'Sell Package'), ('renew', 'Renew Package'),
    ('checkin', 'Front-desk Check-in'), ('check-in', 'Front-desk Check-in'),
    ('assignment', 'PT Assignment'), ('assign', 'PT Assignment'),
    ('workout', 'My Workout Plan'), ('note', 'Trainer Notes'),
    ('progress', 'My Progress'), ('360', 'Member 360 Profile (Admin)'),
    ('nutrition', 'Meal Journal (+ AI Scan)'), ('meal', 'Meal Journal (+ AI Scan)'),
    ('food', 'Meal Journal (+ AI Scan)'), ('calorie', 'Calorie Summary'),
    ('barcode', 'Meal Journal (+ AI Scan)'),
    ('dashboard', 'Admin Dashboard'), ('audit', 'Audit Logs'),
    ('notification', 'Notifications (Admin)'),
    ('auth', 'User Login'), ('login', 'User Login'), ('register', 'User Register'),
    ('password', 'Reset Password (OTP)'), ('otp', 'Reset Password (OTP)'),
    ('account', 'My Profile'), ('profile', 'My Profile'),
    ('trainer', 'Manage PTs'), ('member', 'Manage Members'), ('staff', 'Manage Staff Accounts'),
    ('user', 'Manage User Accounts'),
]

TYPE2LABEL = {
    'feat': 'Task', 'fix': 'Defect', 'docs': 'Task', 'chore': 'Task',
    'test': 'Task', 'refactor': 'Task', 'ci': 'Task', 'style': 'Task',
}


def iter_of(d):
    return min(max((d - START).days // 14 + 1, 1), 4)


def due_of(it):
    return START + timedelta(days=14 * it - 1)


def fn_of(subject, product):
    s = subject.lower()
    for k, v in KEY2FN:
        if k in s and v in product:
            return v
    return ''


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('template')
    ap.add_argument('out')
    ap.add_argument('tracking')
    a = ap.parse_args()

    # ten function hop le
    tw = load_workbook(a.tracking, data_only=True)
    sheet = next(s for s in tw.sheetnames if s.lower() in ('product', 'project'))
    ws = tw[sheet]
    hrow = next(r for r in range(1, 8)
                if any(str(c.value or '').lower().startswith('screen') for c in ws[r]))
    ci = next(i for i, c in enumerate(ws[hrow])
              if str(c.value or '').lower().startswith('screen'))
    product = {str(r[ci].value).strip() for r in ws.iter_rows(min_row=hrow + 1)
               if r[ci].value}

    rows = []
    for repo_name, path, url in REPOS:
        raw = subprocess.run(
            ['git', 'log', '--reverse', '--date=short',
             '--format=%H|%ad|%an|%s'],
            cwd=path, capture_output=True, text=True,
            encoding='utf-8', errors='replace').stdout
        for ln in raw.splitlines():
            parts = ln.split('|', 3)
            if len(parts) < 4:
                continue
            sha, ds, an, subj = parts
            try:
                d = date.fromisoformat(ds.strip())
            except ValueError:
                continue
            if subj.lower().startswith('merge '):
                continue
            m = re.match(r'^(\w+)(?:\([^)]*\))?!?:\s*(.+)$', subj)
            typ = m.group(1).lower() if m else ''
            title = m.group(2) if m else subj
            label = TYPE2LABEL.get(typ, 'Task')
            it = iter_of(d)
            rows.append({
                'title': title[:120],
                'desc': '[{}] {}'.format(repo_name, subj),
                'url': '{}/commit/{}'.format(url, sha),
                'state': 'Closed',
                'assignee': an,
                'created': d.isoformat(),
                'due': due_of(it).isoformat(),
                'milestone': 'iter{}'.format(it),
                'labels': '{}, 3_Done'.format(label),
                'fn': fn_of(subj, product),
            })

    shutil.copy(a.template, a.out)
    wb = load_workbook(a.out)
    s = wb[wb.sheetnames[0]]
    hdr = {str(c.value).strip(): i for i, c in enumerate(s[1]) if c.value}

    def col(*names):
        for n in names:
            for k, i in hdr.items():
                if k.lower().startswith(n.lower()):
                    return i
        return None

    C = {k: col(*v) for k, v in {
        'title': ('Title',), 'desc': ('Description',), 'id': ('Issue ID',),
        'url': ('URL',), 'state': ('State',), 'assignee': ('Assignee',),
        'created': ('Created At',), 'due': ('Due Date',),
        'milestone': ('Milestone',), 'labels': ('Labels',),
        'fn': ('Functions/Screens', 'Function'),
    }.items()}

    s.delete_rows(2, max(s.max_row - 1, 1))
    for i, r in enumerate(rows, 1):
        rr = i + 1

        def put(k, v):
            if C.get(k) is not None:
                s.cell(row=rr, column=C[k] + 1, value=v)
        put('id', i)
        for k in ('title', 'desc', 'url', 'state', 'assignee', 'created',
                  'due', 'milestone', 'labels', 'fn'):
            put(k, r[k])
    wb.save(a.out)

    import collections
    print('Da ghi {} issue tu git log that -> {}'.format(len(rows), a.out))
    print('  theo milestone:', dict(sorted(
        collections.Counter(r['milestone'] for r in rows).items())))
    print('  theo label    :', dict(collections.Counter(
        r['labels'].split(',')[0] for r in rows).most_common()))
    no_fn = sum(1 for r in rows if not r['fn'])
    print('  map duoc Functions/Screens: {}/{}'.format(len(rows) - no_fn, len(rows)))
    if no_fn:
        print('  ({} issue khong map duoc -> de trong, can dien tay)'.format(no_fn))


if __name__ == '__main__':
    main()
