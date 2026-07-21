"""Dien cot In Charge cua Project Tracking theo PHAN CONG NHOM DA CHOT.

Nguon su that: bang NHOM trong `gen_phancong.py` (sinh ra docs/PHAN_CONG.md),
la phan cong nhom da thong nhat 2026-07-20.

LICH SU: ban dau script nay UOC LUONG tu so commit git ("ai commit nhieu nhat
trong thu muc feature = chu feature"). Cach do SAI HAN so voi phan cong that:
    do tu git      : anhdaijka 30 · BanhMiChao 15 · Loc-LX 11 · vandam2005 2 · Minhdicodedao 1
    nhom da chot   : BanhMiChao 12 · anhdaijka 7 · Loc-LX 8 · vandam2005 9 · Minhdicodedao 10
Nguyen nhan: anhdaijka dung ra dung khung FE cho gan het man trong tuan dau, nen
dem commit thi bon nguoi kia gan nhu trang — trong khi ho moi la nguoi lam
nghiep vu cua man do. Cot nay quyet dinh Individual Results (60% diem), de nham
la vandam2005/Minhdicodedao mat gan het diem ca nhan.

=> Khong doan tu git nua. Doc thang phan cong nhom.

Dung: python fill_in_charge.py <tracking.xlsx> <inventory.csv> [--add-backend]
"""
import argparse
import ast
import collections
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Thieu openpyxl. Chay: uv run --with openpyxl python ' + __file__)

sys.path.insert(0, str(Path(__file__).parent))
from fill_tracking import NAMES


def doc_phan_cong():
    """Doc bang NHOM tu gen_phancong.py.

    Khong `import gen_phancong` duoc: file do GHI docs/PHAN_CONG.md ngay luc
    import (code chay o muc module). Nen parse bang AST va literal_eval —
    NHOM chi gom dict/list/str nen an toan.
    """
    src = (Path(__file__).parent / 'gen_phancong.py').read_text(encoding='utf-8')
    tree = ast.parse(src)
    for node in tree.body:
        if not (isinstance(node, ast.Assign) and any(
                getattr(t, 'id', '') == 'NHOM' for t in node.targets)):
            continue
        # Moi phan tu la `dict(id=..., git=..., screens=[...], be=[...])` — la
        # ast.Call chu khong phai dict literal, nen literal_eval ca cum se nem
        # loi. Chi boc dung 3 khoa can dung, gia tri deu la literal.
        out = []
        for el in node.value.elts:
            g = {}
            for kw in el.keywords:
                if kw.arg in ('git', 'screens', 'be', 'nonui'):
                    g[kw.arg] = ast.literal_eval(kw.value)
            out.append(g)
        return out
    sys.exit('Khong tim thay bang NHOM trong gen_phancong.py')


# 3 function khong co man hinh trong Tracking <- muc `nonui` cua phan cong.
# Ten trong Tracking do fill_tracking.py dat, khong trung chu voi mo ta ben
# gen_phancong nen phai map bang tu khoa.
NONUI_KEY = [
    ('VNPay IPN', 'ipn'),
    ('Auto-Cancel', 'auto-cancel'),
    ('Lazy Expire', 'lazy-expire'),
]


def bang_phan_cong():
    """-> (route -> git, feature backend -> git, ten non-UI -> git)"""
    man, be, nonui = {}, {}, {}
    for g in doc_phan_cong():
        for s in g.get('screens', []):
            man[s] = g['git']
        for f in g.get('be', []):
            be[f] = g['git']
        for mo_ta in g.get('nonui', []):
            low = mo_ta.lower()
            for ten, khoa in NONUI_KEY:
                if khoa in low:
                    nonui[ten] = g['git']
    return man, be, nonui

BE_DESC = {
    'Auth': ('Authentication', 'API dang nhap, dang ky, quen mat khau (OTP), doi mat khau, '
                               'Google login, refresh/revoke token'),
    'Users': ('User Management', 'API quan ly tai khoan: tao, khoa/mo, phan quyen, reset mat khau'),
    'Members': ('Member Management', 'API ho so hoi vien: tao, tim kiem, cap nhat, Member 360'),
    'Trainers': ('Trainer Management', 'API ho so PT: tao, danh sach, cap nhat'),
    'Dashboard': ('Dashboard & Audit', 'API doanh thu, thong ke van hanh, audit log'),
    'Billing': ('Membership & Billing', 'API goi tap, ban/gia han, thanh toan, VNPay (tao URL, IPN, return)'),
    'Nutrition': ('Nutrition', 'API bua an, muc tieu calo, danh muc mon, quet anh AI (Gemini)'),
    'Training': ('PT Training', 'API phan cong PT, giao an, ghi chu, tien do'),
    'CheckIns': ('Check-in', 'API check-in cua member/le tan/PT, gioi han 2 luot/ngay'),
    'Account': ('Account', 'API ho so ca nhan, doi avatar (Cloudinary)'),
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('tracking')
    ap.add_argument('inventory')
    ap.add_argument('--add-backend', action='store_true')
    a = ap.parse_args()

    man_owner, be_owner, nonui_owner = bang_phan_cong()

    with open(a.inventory, encoding='utf-8-sig') as fh:
        inv = [r for r in csv.DictReader(fh) if r['kind'] == 'Screen']

    wb = load_workbook(a.tracking)
    sheet = next(s for s in wb.sheetnames if s.lower() in ('product', 'project'))
    ws = wb[sheet]
    hrow = next(r for r in range(1, 8)
                if any(str(c.value or '').lower().startswith('screen') for c in ws[r]))
    hdr = {str(c.value).strip(): i for i, c in enumerate(ws[hrow]) if c.value}

    def col(*names):
        for n in names:
            for k, i in hdr.items():
                if k.lower().startswith(n.lower()):
                    return i
        return None

    c_no, c_fn = col('#'), col('Screen/Function', 'Screen')
    c_ft, c_ac = col('Feature'), col('Actor')
    c_de, c_ic = col('Screen/Function Description', 'Description'), col('In Charge')
    c_st, c_acl = col('Status'), col('Actual')

    # ten man -> chu
    route_of = {NAMES.get(r['name'], r['name']): r['name'] for r in inv}
    stats = collections.Counter()
    thieu = []          # man co trong inventory ma phan cong chua nhac toi
    last = hrow
    for row in ws.iter_rows(min_row=hrow + 1):
        nm = row[c_fn].value
        if not nm:
            continue
        last = row[0].row
        ten = str(nm).strip()
        route = route_of.get(ten)
        if not route:
            # Khong phai man hinh -> co the la 1 trong 3 function non-UI.
            # Truoc day nhanh nay `continue` thang, nen 3 dong do giu nguyen gia
            # tri mac dinh cua fill_tracking (--in-charge) va bi gan nham nguoi.
            o = next((v for k, v in nonui_owner.items() if k.lower() in ten.lower()), '')
            if o and c_ic is not None:
                row[c_ic].value = o
                stats[o] += 1
            elif 'API' not in ten:
                thieu.append(ten)
            continue
        o = man_owner.get(route, '')
        if not o:
            thieu.append(route)
        if o and c_ic is not None:
            row[c_ic].value = o
            stats[o] += 1

    if a.add_backend:
        n = last - hrow
        for name, owner in ((k, be_owner.get(k, '')) for k in sorted(BE_DESC)):
            n += 1
            last += 1
            feat, desc = BE_DESC[name]

            def put(ci, v):
                if ci is not None:
                    ws.cell(row=last, column=ci + 1, value=v)
            put(c_no, n)
            put(c_fn, '{} API'.format(name))
            put(c_ft, feat)
            put(c_ac, 'System')
            put(c_de, desc)
            put(c_ic, owner)
            put(c_st, 'Done')
            put(c_acl, 'iter1')
            stats[owner] += 1

    wb.save(a.tracking)
    print('Da dien In Charge -> {}'.format(a.tracking))
    print('Nguon: bang NHOM trong gen_phancong.py (= docs/PHAN_CONG.md).')
    print('\nSo dong moi nguoi phu trach:')
    for k, v in stats.most_common():
        print('  {:<16} {:>2} dong'.format(k or '(khong ro)', v))
    if thieu:
        print('\nCANH BAO: {} man co trong inventory nhung PHAN CONG CHUA NHAC TOI '
              '-> de trong:'.format(len(thieu)))
        for r in thieu:
            print('  ', r)
        print('  Sua bang NHOM trong gen_phancong.py roi chay lai.')


if __name__ == '__main__':
    main()
